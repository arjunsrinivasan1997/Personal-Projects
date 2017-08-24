import smtplib
import sys
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from string import Template

from openpyxl import load_workbook

'Basic info. Change this to change how program runs'
numberOfEmailsToBeSent = 1
startRow = 10
senderAddress = ""
senderPassword = ""
senderName = ""

'Check to see if command line arguments were given'
if len(sys.argv) >= 2:
    try:
        numberOfEmailsToBeSent = int(sys.argv[1])
    except ValueError:
        raise ValueError("Number of Emails to be sent must be an integer")
if len(sys.argv) == 3:
    try:
        startRow = int(sys.argv[2])
    except ValueError:
        raise ValueError("Start Row must be an integer")

'Function to read in file and create a template object from that file'


def read_template(filename):
    with open(filename, 'r', encoding='utf-8') as template_file:
        template_file_content = template_file.read()
    return Template(template_file_content)


"""Removes the unnecessary space at the end of the message, and removes the unnecessary 
   period at the end of a string"""


def clean_string(message):
    if message[0] == " ":
        return clean_string(message[1:len(message)])
    if message[len(message) - 1] == " ":
        return clean_string(message[0:len(message) - 1])
    if "." == message[len(message) - 1]:
        return clean_string(message[0:len(message) - 1])
    return message


'Setting Up Email Server'

s = smtplib.SMTP(host='smtp.gmail.com', port=587)
s.starttls()
s.login(senderAddress, senderPassword)

'Setting Up Excel Sheet'
wb = load_workbook('Outreach Track Sheet.xlsx')
ws = wb["Master List USA"]
currRow = startRow
template = read_template('template.txt')
"""Looping through each row specified by startRow and numberOfEmailsToBeSent, gathering a name, email, media type and 
   personal message, adding those values to the template, and sending the email"""
for i in range(numberOfEmailsToBeSent):
    msg = MIMEMultipart()
    try:
        name = clean_string(ws.cell(row=currRow, column=1).value)
        email = clean_string(ws.cell(row=currRow, column=2).value)
        mediaType = clean_string(ws.cell(row=currRow, column=4).value)
        personalMessage = clean_string(ws.cell(row=currRow, column=5).value)
        sendEmail = True
    except TypeError:
        sendEmail = False

    if name is None:
        sendEmail = False
    if email is None:
        sendEmail = False
    if mediaType is None:
        sendEmail = False
    if personalMessage is None:
        sendEmail = False
    if sendEmail:
        body = template.substitute(PERSON_NAME=name, TYPE_OF_MEDIA=mediaType, Personal_Message=personalMessage,
                                   Sender_Name=senderName)
        msg['From'] = senderAddress
        msg['To'] = email
        msg['Subject'] = "Cuddle Cub"
        msg.attach(MIMEText(body, 'plain'))

        s.send_message(msg)
        print("Sent email to " + str(name) + " at " + str(email))
        ws.cell(row=currRow, column=6).value = senderName

        dateCell = ws.cell(row=currRow, column=7)
        dateCell.value = date.today()
        dateCell.number_format = "MM/DD/YYYY"

        del msg
    else:
        print("ERROR: Could not send email to " + str(name) + " at " + str(email) +
              ". Check row " + str(currRow) + " and ensure none of the first four columns are blank")
    currRow += 1
wb.save('Outreach Track Sheet.xlsx')
print("Finished sending emails")
