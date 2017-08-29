from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver


class Analyzer:
    """Class that analyzes links"""
    mainLink = 'http://mha1.nic.in/ipr/ips_ersheet_new/home.aspx'

    def __init__(self, year, cadre):
        self.year = year
        self.cadre = cadre
        self.links = []
        self.values = {}
        self.officers = []
        self.optionsMap = {}
        self.driver = webdriver.Chrome(executable_path='Driver/chromedriver')
        self.driver.get(self.mainLink)





    def analyze(self):
        for year in self.year:
            for cadre in self.cadre:
                form = self.driver.find_element_by_id('form1')
                self.all_options = form.find_elements_by_tag_name("option")
                self.parser = BeautifulSoup(self.driver.page_source, "html5lib")

                cadreSelect = self.findElement("select", "id", "ddlCadre")
                for option in cadreSelect.children:
                    if option.string != "\n\t":
                        try:
                            self.optionsMap[option.string] = option.attrs["value"]
                        except AttributeError:
                            pass
                for option in self.all_options:
                    self.values[option.get_attribute("value")] = option
                self.values[str(year)].click()
                self.values[self.optionsMap[str(cadre)]].click()
                self.driver.find_element_by_id("btnSubmit").click()
                self.links = self.driver.find_elements_by_xpath("//a[@href]")
                self.newLinks = self.driver.find_elements_by_xpath("//a[@href]")
                count = 0
                for link in self.links:
                    if self.newLinks[count].text == "View ER SHEET":
                        if count == 1:
                            self.links[count].click()
                        else:
                            self.newLinks[count].click()
                        self.parseLink()
                        self.driver.back()
                        self.newLinks = self.driver.find_elements_by_xpath("//a[@href]")
                    count += 1
        self.driver.quit()

    def parseLink(self):
        self.parser = BeautifulSoup(self.driver.page_source, "html5lib")
        'Gathering Basic info'
        officer = Officer(self.findElement('span', 'id', 'lblName').text,
                          self.findElement('span', 'id', 'lblDOB').text,
                          self.findElement('span', 'id', 'lblCadre').text,
                          self.findElement('span', 'id', 'lblBatch').text,
                          self.findElement('span', 'id', 'lblIdentityNo').text,
                          self.findElement('span', 'id', 'lblDateOfAppointment').text,
                          self.findElement('span', 'id', 'lblSourceOfRecruitment').text)
        'Gathering Education Info Each row in the education table has 10 elements in it'
        educationInfo = list(self.findElement('table', 'id', 'gvEducation').find_all('td'))
        education = []
        for i in range(int(len(educationInfo) / 10)):
            tempEduRow = EducationRow(educationInfo[i * 10 + 1].text, educationInfo[i * 10 + 2].text,
                                      educationInfo[i * 10 + 3].text, educationInfo[i * 10 + 4].text,
                                      educationInfo[i * 10 + 5].text, educationInfo[i * 10 + 6].text,
                                      educationInfo[i * 10 + 7].text, educationInfo[i * 10 + 8].text,
                                      educationInfo[i * 10 + 9].text)
            education.append(tempEduRow)
        officer.education = education
        postings = []
        postingInfo = list(self.findElement('table', 'id', 'gvPostingExperience').find_all('td'))
        for i in range(int(len(postingInfo) / 10)):
            tempPostRow = PostingRow(postingInfo[i * 10 + 0].text.strip(), postingInfo[i * 10 + 1].text,
                                     postingInfo[i * 10 + 2].text,
                                     postingInfo[i * 10 + 3].text, postingInfo[i * 10 + 4].text,
                                     postingInfo[i * 10 + 5].text, postingInfo[i * 10 + 6].text,
                                     postingInfo[i * 10 + 7].text, postingInfo[i * 10 + 8].text,
                                     postingInfo[i * 10 + 9].text)
            postings.append(tempPostRow)
        officer.postings = postings
        training = []
        domesticTraining = list(self.findElement('table', 'id', 'gvDomesticTraining').find_all('td'))
        for i in range(int(len(domesticTraining) / 6)):
            tempTrainingRow = TrainingRow(domesticTraining[i * 6 + 0].text.strip(), domesticTraining[i * 6 + 1].text,
                                          domesticTraining[i * 6 + 2].text,
                                          domesticTraining[i * 6 + 3].text, domesticTraining[i * 6 + 4].text,
                                          domesticTraining[i * 6 + 5].text, "Domestic")
            training.append(tempTrainingRow)
        foreignTraining = list(self.findElement('table', 'id', 'gvForeignTraining').find_all('td'))
        for i in range(int(len(foreignTraining) / 6)):
            tempTrainingRow = TrainingRow(foreignTraining[i * 6 + 0].text.strip(), foreignTraining[i * 6 + 1].text,
                                          foreignTraining[i * 6 + 2].text,
                                          foreignTraining[i * 6 + 3].text, foreignTraining[i * 6 + 4].text,
                                          foreignTraining[i * 6 + 5].text, "Foreign")
            training.append(tempTrainingRow)
        officer.training = training

        medalsTable = list(self.findElement('table', 'id', 'gvMedal').find_all('td'))
        medals = []
        for i in range(int(len(medalsTable) / 3)):
            tempMedalRow = MedalRow(medalsTable[i * 3 + 0].string.strip(), medalsTable[i * 3 + 1].string.rstrip(),
                                    medalsTable[i * 3 + 2].string.rstrip())
            medals.append(tempMedalRow)
        officer.medals = medals
        self.officers.append(officer)

    def findElement(self, tag, atrKey, atrValue):
        return self.parser.find(str(tag), {str(atrKey): str(atrValue)})


class EducationRow:
    def __init__(self, qualification, discipline, spec1, spec2, year, division, institution, university, place):
        self.qualification = qualification
        self.discipline = discipline
        if spec1 == "-":
            self.spec1 = ""
        else:
            self.spec1 = spec1
        if spec2 == "-":
            self.spec2 = ""
        else:
            self.spec2 = spec2
        self.year = year
        self.division = division
        self.institution = institution
        self.university = university
        self.place = place

    def __str__(self):
        s = ""
        s += ("Qualification:" + str(self.qualification) + " ")
        s += ("Discipline:" + str(self.discipline) + " ")
        s += ("spec1:" + str(self.spec1) + " ")
        s += ("spec2:" + str(self.spec2) + " ")
        s += ("Year:" + str(self.year) + " ")
        s += ("Division:" + str(self.division) + " ")
        s += ("Institution" + str(self.institution) + " ")
        s += ("University" + str(self.university) + " ")
        s += ("Place:" + str(self.place) + " ")
        return s

    def fillData(self, row, sheet, startColumn):
        sheet.cell(row=row, column=startColumn + 1).value = self.qualification
        sheet.cell(row=row, column=startColumn + 2).value = self.year
        sheet.cell(row=row, column=startColumn + 3).value = self.institution
        sheet.cell(row=row, column=startColumn + 4).value = self.university
        sheet.cell(row=row, column=startColumn + 5).value = self.place
        sheet.cell(row=row, column=startColumn + 6).value = self.discipline
        sheet.cell(row=row, column=startColumn + 7).value = self.spec1
        sheet.cell(row=row, column=startColumn + 8).value = self.spec2


class PostingRow():
    def __init__(self, number, type, level, designation, ministry, place, experienceMajor, experienceMinor, startDate,
                 endDate, ):
        self.number = number
        self.type = type
        self.level = level
        self.designation = designation
        self.ministry = ministry
        self.place = place
        self.experienceMajor = experienceMajor
        self.experienceMinor = experienceMinor
        self.startDate = startDate
        self.endDate = endDate

    def __str__(self):
        s = ""
        s += ("Type:" + str(self.type) + " ")
        s += ("Level:" + str(self.level) + " ")
        s += ("Designation:" + str(self.designation) + " ")
        s += ("Ministry:" + str(self.ministry) + " ")
        s += ("Place:" + str(self.place) + " ")
        s += ("Experience Major:" + str(self.experienceMajor) + " ")
        s += ("Experience Minor:" + str(self.experienceMinor) + " ")
        s += ("Start Date:" + str(self.startDate) + " ")
        s += ("End Date:" + str(self.endDate) + " ")
        return s

    def fillData(self, row, sheet, startColumn):
        sheet.cell(row=row, column=startColumn + 1).value = self.number
        sheet.cell(row=row, column=startColumn + 2).value = self.type
        sheet.cell(row=row, column=startColumn + 3).value = self.level
        sheet.cell(row=row, column=startColumn + 4).value = self.designation
        sheet.cell(row=row, column=startColumn + 5).value = self.ministry
        sheet.cell(row=row, column=startColumn + 6).value = self.place
        sheet.cell(row=row, column=startColumn + 7).value = self.experienceMajor
        sheet.cell(row=row, column=startColumn + 8).value = self.experienceMinor
        sheet.cell(row=row, column=startColumn + 9).value = self.startDate
        sheet.cell(row=row, column=startColumn + 10).value = self.endDate


class MedalRow():
    def __init__(self, medalNumber, medalDescription, year):
        self.medalNumber = medalNumber
        self.medalDescription = medalDescription
        self.year = year

    def __str__(self):
        s = ""
        s += "Medal Number : " + str(self.medalNumber) + " "
        s += "Medal Description: " + str(self.medalDescription) + " "
        s += "Year: " + str(self.year) + " "
        return s
    def fillData(self,row,sheet,startColumn):
        sheet.cell(row=row, column=startColumn + 1).value = self.medalNumber
        sheet.cell(row=row, column=startColumn + 2).value = self.medalDescription
        sheet.cell(row=row, column=startColumn + 3).value = self.year



class TrainingRow():
    def __init__(self, number, courseName, courseTitle, place, startDate, endDate, type):
        self.number = number
        self.courseName = courseName
        self.courseTitle = courseTitle
        self.place = place
        self.startDate = startDate
        self.endDate = endDate
        self.type = type

    def __str__(self):
        s = ""
        s += ("Number:" + str(self.number) + " ")
        s += ("Course Name:" + str(self.courseName) + " ")
        s += ("Course Title:" + str(self.courseTitle) + " ")
        s += ("Place:" + str(self.place) + " ")
        s += ("Start Date:" + str(self.startDate) + " ")
        s += ("End Date:" + str(self.endDate) + " ")
        s += ("Type:" + str(self.type) + " ")
        return s

    def fillData(self, row, sheet, startColumn):
        sheet.cell(row=row, column=startColumn + 1).value = self.number
        sheet.cell(row=row, column=startColumn + 2).value = self.type
        sheet.cell(row=row, column=startColumn + 3).value = self.courseName
        sheet.cell(row=row, column=startColumn + 4).value = self.place
        sheet.cell(row=row, column=startColumn + 5).value = self.startDate
        sheet.cell(row=row, column=startColumn + 6).value = self.endDate


class Officer():
    def __init__(self, name, dob, cadre, batch, idNumber, dateOfAppointment, sourceOfRecuritment, ):
        if name[0:4] == 'Shri':
            self.gender = 'Shri'
            self.name = name[5:len(name)]
        elif name[0:4] == 'Smt.':
            self.gender = "Smt."
            self.name = name[5:len(name)]
        else:
            self.gender = ''
            self.name = name
        date = dob.split("/")
        self.dob = date[1] + "/" + date[0] + "/" + date[2]
        self.cadre = cadre
        self.batch = batch
        self.idNumber = idNumber
        self.dateOfAppointment = dateOfAppointment
        self.sourceOfRecruitment = sourceOfRecuritment
        self.education = []
        self.postings = []
        self.training = []
        self.medals = []

    def __str__(self):
        s = ""
        s += "Gender:" + str(self.gender) + "\n"
        s += "Name:" + str(self.name) + "\n"
        s += "Date of birth:" + str(self.dob) + "\n"
        s += "Cadre:" + str(self.cadre) + "\n"
        s += "Batch:" + str(self.batch) + "\n"
        s += "Id Number:" + str(self.idNumber) + "\n"
        s += "Appointment Date:" + str(self.dateOfAppointment) + "\n"
        s += "Recruitment Source:" + str(self.sourceOfRecruitment) + "\n"
        s += "Education" + "\n"
        for educationRow in self.education:
            s += str(educationRow) + "\n"
        s += "Training" + "\n"
        for trainingRow in self.training:
            s += str(trainingRow) + "\n"
        s += "Postings" + "\n"
        for postingsRow in self.postings:
            s += str(postingsRow) + "\n"
        s += "Medals" + "\n"
        for medalRow in self.medals:
            s += str(medalRow) + "\n"
        return s

    def fillData(self, educationStartRow, trainingStartRow, workbook):
        rows = []
        eduRow = educationStartRow
        trainingRow = trainingStartRow
        educationSheet = workbook['Education']
        trainingSheet = workbook['Training and Posting info']
        for education in self.education:
            self.fillHeader(eduRow, educationSheet)
            education.fillData(eduRow, educationSheet, 4)
            eduRow += 1
        for training in self.training:
            self.fillHeader(trainingRow, trainingSheet)
            training.fillData(trainingRow, trainingSheet, 4)
            trainingRow += 1
        for posting in self.postings:
            self.fillHeader(trainingRow,trainingSheet)
            posting.fillData(trainingRow,trainingSheet,10)
            trainingRow += 1
        for medals in self.medals:
            self.fillHeader(trainingRow,trainingSheet)
            medals.fillData(trainingRow,trainingSheet,20)
            trainingRow += 1
        rows.append(eduRow)
        rows.append(trainingRow)

        return rows

    def fillHeader(self, row, sheet):
        sheet.cell(row=row, column=1).value = self.gender
        sheet.cell(row=row,column=2).value = self.name
        sheet.cell(row=row, column=3).value = self.dob
        sheet.cell(row=row, column=4).value = self.sourceOfRecruitment


years = [1988,1989,1990,1991]
cadres = ['Kerala']
analayze = Analyzer(years, cadres)
analayze.analyze()
dataWorkbook = load_workbook("Template.xlsx")
filename = "test" + ".xlsx"
dataWorkbook.save(filename)
educationRow = 2
trainingRow = 2

for officer in analayze.officers:
    newRows = officer.fillData(educationRow, trainingRow, dataWorkbook)
    educationRow = newRows[0]
    trainingRow = newRows[1]
    dataWorkbook.save(filename)

    'print(officer)'

